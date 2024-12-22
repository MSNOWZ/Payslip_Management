import time #importing time library
import tkinter.messagebox #importing tkinter library to use in GUI
from tkinter import *
import xlrd #importing xlrd library
import openpyxl #importing onepyxl library to use excel sheets
from openpyxl import Workbook
import pathlib

'''Drawing the frame used to display all the GUI elements'''
root = Tk()
root.title("Employee payroll system") #naming the window
root.geometry('1350x650+0+0') #size of the window
root.configure(background="grey20")

Tops = Frame(root, width=1350, height=50, bd=8, bg="grey20")
Tops.pack(side=TOP)

f1 = Frame(root, width=600, height=600, bd=8, bg="grey20")
f1.pack(side=LEFT)
f2 = Frame(root, width=300, height=700, bd=8, bg="grey20")
f2.pack(side=RIGHT)

fla = Frame(f1, width=600, height=200, bd=8, bg="grey20")
fla.pack(side=TOP)
flb = Frame(f1, width=300, height=600, bd=8, bg="grey20")
flb.pack(side=TOP)

lblinfo = Label(Tops, font=('Times', 45, 'bold'), text="Employee Payment Management system ", bd=2, fg="white", bg="grey20")
lblinfo.grid(row=0, column=0)

'''method used to check if the file exists, and create one if it doesnt exist'''
file = pathlib.Path("employees.xlsx")
if file.exists():
  pass #exits the loop and continues the program
else:
  file = Workbook()
  sheet = file.active
  sheet["A1"] = "Name"
  sheet["B1"] = "Address"
  sheet["C1"] = "Employer"
  sheet["D1"] = "Hours Worked"
  sheet["E1"] = "Hourly Wage"
  sheet["F1"] = "Employee IDe"
  sheet["G1"] = "Payable"
  sheet["H1"] = "Tax"
  sheet["I1"] = "Net pay"
  sheet["J1"] = "Overtime"
  file.save("employees.xlsx")

'''method to export the payslip information to an excel file'''
def export_to_excel():
  name = Name.get() #assigning variables the values in the text boxes
  address = Address.get()
  employer = Employer.get()
  hours = HoursWorked.get()
  hourlywage = wageshour.get()
  eid = EID.get()
  payable = Payable.get()
  tax = Taxable.get()
  netpay = NetPayable.get()
  overtime = OverTimeBonus.get()
  file = openpyxl.load_workbook("employees.xlsx")
  sheet = file.active
  sheet.cell(column=1, row=sheet.max_row+1, value=name)
  sheet.cell(column=2, row=sheet.max_row, value=address)
  sheet.cell(column=3, row=sheet.max_row, value=employer)
  sheet.cell(column=4, row=sheet.max_row, value=hours)
  sheet.cell(column=5, row=sheet.max_row, value=hourlywage)
  sheet.cell(column=6, row=sheet.max_row, value=eid)
  sheet.cell(column=7, row=sheet.max_row, value=payable)
  sheet.cell(column=8, row=sheet.max_row, value=tax)
  sheet.cell(column=9, row=sheet.max_row, value=netpay)
  sheet.cell(column=10, row=sheet.max_row, value=overtime)
  file.save("employees.xlsx")

'''method to be used with the exit button on the GUI'''
def exit():
  exit = tkinter.messagebox.askyesno("Employee system", "Do you want to exit the system")
  if exit == 1:
    root.destroy() #destroys the window and exits the program
    return

'''method to reset the text bars so new information can be entered'''
def reset(): #assign empty strings to each texbox
  Name.set("")
  Address.set("")
  HoursWorked.set("")
  wageshour.set("")
  Payable.set("")
  Taxable.set("")
  NetPayable.set("")
  GrossPayable.set("")
  OverTimeBonus.set("")
  Employer.set("")
  EID.set("")
  txtpayslip.delete("1.0",END)

'''method to enter the information in the text boxes to the payslip'''
def enterinfo():
  txtpayslip.delete("1.0",END)
  txtpayslip.insert(END, "\t\tPay Slip\n\n")
  txtpayslip.insert(END, "Name :\t\t"+Name.get()+"\n\n")
  txtpayslip.insert(END, "Address :\t\t"+Address.get()+"\n\n")
  txtpayslip.insert(END, "Employer :\t\t"+Employer.get()+"\n\n")
  txtpayslip.insert(END, "Employee ID :\t\t"+EID.get()+"\n\n")
  txtpayslip.insert(END, "Hours Worked :\t\t"+HoursWorked.get()+"\n\n")
  txtpayslip.insert(END, "Net Payable :\t\t"+NetPayable.get()+"\n\n")
  txtpayslip.insert(END, "Wages per hour :\t\t"+wageshour.get()+"\n\n")
  txtpayslip.insert(END, "Tax Paid :\t\t"+Taxable.get()+"\n\n")
  txtpayslip.insert(END, "Payable :\t\t"+Payable.get()+"\n\n")

'''method to calculate weeklywages and show the output'''
def weeklywages():
  txtpayslip.delete("1.0", END)
  hoursworkedperweek = float(HoursWorked.get())
  wagesperhours = float(wageshour.get())

  paydue = wagesperhours*hoursworkedperweek
  paymentdue = "AED",str('%.2f'%(paydue))
  Payable.set(paymentdue)

  tax = paydue*0.2
  taxable = "AED", str('%.2f'%(tax))
  Taxable.set(taxable)

  netpay = paydue-tax
  netpays = "AED", str('%.2f'%(netpay))
  NetPayable.set(netpays)

  if hoursworkedperweek > 40:
    overtimehours = (hoursworkedperweek-40)+wagesperhours*1.5
    overtime = "AED", str('%.2f'%(overtimehours))
    OverTimeBonus.set(overtime)
  elif hoursworkedperweek <= 40:
    overtimepay = (hoursworkedperweek-40)+wagesperhours*1.5
    overtimehrs = "AED", str('%.2f'%(overtimepay))
    OverTimeBonus.set(overtimehrs)
  return

'''assigning variables'''
Name = StringVar()
Address = StringVar()
HoursWorked = StringVar()
wageshour = StringVar()
Payable = StringVar()
Taxable = StringVar()
NetPayable = StringVar()
GrossPayable = StringVar()
OverTimeBonus = StringVar()
Employer = StringVar()
EID = StringVar()
TimeOfOrder = StringVar()
DateOfOrder = StringVar()

DateOfOrder.set(time.strftime("%d/%m/%Y")) #using the time library to set time for the payslip

'''Labels used in the GUI with the text boxes'''
Label(fla, text="Name", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=0,column=0)
Label(fla, text="Address", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=0,column=2)
Label(fla, text="Employer", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=1,column=0)
Label(fla, text="Employee ID", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=1,column=2)
Label(fla, text="Hours Worked", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=2,column=0)
Label(fla, text="Hourly Rate", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=2,column=2)
Label(fla, text="Tax", font=('arial', 16, 'bold'), bd=2, anchor='w', fg="white", bg="gray20").grid(row=3,column=0)
Label(fla, text="OverTime", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=3,column=2)
Label(fla, text="GrossPay", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=4, column=0)
Label(fla, text="Net Pay", font=('arial', 16, 'bold'), bd=2, fg="white", bg="grey20").grid(row=4, column=2)

'''Text boxes used in the GUI'''
etxname = Entry(fla, textvariable=Name, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxname.grid(row=0, column=1)

etxaddress = Entry(fla, textvariable=Address, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxaddress.grid(row=0, column=3)

etxemployer = Entry(fla, textvariable=Employer, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxemployer.grid(row=1, column=1)

etxhoursworked = Entry(fla, textvariable=HoursWorked, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxhoursworked.grid(row=2, column=1)

etxwagesperhours = Entry(fla, textvariable=wageshour, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxwagesperhours.grid(row=2, column=3)

etxeid = Entry(fla, textvariable=EID, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxeid.grid(row=1, column=3)

etxgrosspay = Entry(fla, textvariable=Payable, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxgrosspay.grid(row=4, column=1)

etxnetpay = Entry(fla, textvariable=NetPayable, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxnetpay.grid(row=4, column=3)

etxtax = Entry(fla, textvariable=Taxable, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxtax.grid(row=3, column=1)

etxovertime = Entry(fla, textvariable=OverTimeBonus, font=('arial', 16, 'bold'), bd=5, bg="grey75", width=22, justify='left')
etxovertime.grid(row=3, column=3)

'''Widget to display the payslip'''
payslip = Label(f2, textvariable=DateOfOrder, font=('arial', 21, 'bold'), fg="white", bg="grey20").grid(row=0, column=0)
txtpayslip = Text(f2, height=22, width=34, bd=3, font=('arial', 13, 'bold'), fg="black", bg="grey75")
txtpayslip.grid(row=1, column=0)

'''Buttons used in the GUI and functions assigned to each button'''
btnsalary = Button(flb, text='Weekly Salary', padx=10, pady=10, bd=2, font=('arial', 16, 'bold'), width=14, fg="black", bg="grey75", command=weeklywages).grid(row=0, column=0)

btnreset = Button(flb, text='Reset', padx=10, pady=10, bd=2, font=('arial', 16, 'bold'), width=14, command=reset, fg="black", bg="grey75").grid(row=0, column=1)

btnpayslip = Button(flb, text='View Payslip', padx=10, pady=10, bd=2, font=('arial', 16, 'bold'), width=14, command=enterinfo, fg="black", bg="grey75").grid(row=0, column=2)

btnexit = Button(flb, text='Exit System', padx=10, pady=10, bd=2, font=('arial', 16, 'bold'), width=14, command=exit, fg="black", bg="grey75").grid(row=0, column=3)

btnexport = Button(flb, text='Export', padx=10, pady=10, bd=2, font=('arial', 16, 'bold'), width=14, command=export_to_excel, fg="black", bg="grey75").grid(row=0, column=4)

root.mainloop()


